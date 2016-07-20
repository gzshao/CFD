import os
from openpyxl import load_workbook
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.legend_handler import HandlerLine2D


#This python code will read .xlsx file with the following format
#space | name1 |name2 |space|name3 |name4 |....
#space | data11|data21|space|data31|data41|....
#...
#in the end, it will:

#1. plt.plot(name1, name2,'o'), ignoring the first row
#   plt.plot(name3, name4,'o'), ignoring the first row
#   .....
   
#2. use different color schemes

#sample data is shown below


#   Z' (ohm)	-Z'' (ohm)		Z' (ohm)	-Z'' (ohm)
#	6613.33	435.197		7793.73	531.564
#    6655.38	489.891		7858.06	586.49
#    6707.3	567.708		7913.34	675.172
#    6796.18	610.423		8010.48	723.341
#    6777.13	669.022		8009.17	758.735
#    6868.54	738.669		8081.81	861.052
#    6925.87	855.264		8145.29	957.443
#    7021.04	901.222		8296.64	1024.96

os.chdir("/Users/Guozheng/Downloads")
wb = load_workbook(filename='Book3.xlsx')
ws1 = wb['Sheet2']

nRow=ws1.get_highest_row()-1

##### This defines how many figures to plot, thus how many (*2) columns to read ########
nPlot=6
#############

xAxis=np.empty([nRow,nPlot])
yAxis=np.empty([nRow,nPlot])

for k in range(nPlot):
    for i in range(nRow):
        xAxis[i-1,k]=ws1.cell(row=i+1,column=3*k+1).value
        yAxis[i-1,k]=ws1.cell(row=i+1,column=3*k+2).value

line1,=plt.plot(xAxis[:,0],yAxis[:,0],'oc',label='0.5M standard')
line2,=plt.plot(xAxis[:,1],yAxis[:,1],'Dc',label="0.4M standard")
line3,=plt.plot(xAxis[:,2],yAxis[:,2],'dc',label="0.3M standard")
line4,=plt.plot(xAxis[:,3],yAxis[:,3],'pc',label="0.2M standard")
#plt.plot(xAxis[:,4]+2000,yAxis[:,4],'ro')
line5,=plt.plot(xAxis[:,5]+2000,yAxis[:,5],'ro',label="0.5M voltage on 20$\mu$L/min")

#plt.legend(handles=[line1,line2,line3,line4,line5],loc=1)
plt.legend(handler_map={line1:HandlerLine2D(numpoints=1),line2:HandlerLine2D(numpoints=1),line3:HandlerLine2D(numpoints=1),line4:HandlerLine2D(numpoints=1),line5:HandlerLine2D(numpoints=1)})
plt.xlabel("Z'/$\Omega$",fontsize=18)
plt.ylabel("-Z''/$\Omega$",fontsize=18)
#plt.title(r'$\alpha$')
plt.ticklabel_format(style='sci', scilimits=(0,0))
plt.show()